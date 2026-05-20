import streamlit as st
import pandas as pd
from supabase import create_client, Client
from io import BytesIO
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# --- 1. SETUP E CONEXÃO ---
st.set_page_config(page_title="Conciliador Contábil", page_icon="📊", layout="wide")


@st.cache_resource
def init_connection() -> Client:
    if "SUPABASE_URL" not in st.secrets:
        st.error("Faltou configurar `SUPABASE_URL` nos secrets do Streamlit Cloud.")
        st.stop()

    if "SUPABASE_SERVICE_ROLE_KEY" in st.secrets:
        key = st.secrets["SUPABASE_SERVICE_ROLE_KEY"]
    elif "SUPABASE_KEY" in st.secrets:
        key = st.secrets["SUPABASE_KEY"]
    else:
        st.error("Faltou configurar `SUPABASE_KEY` ou `SUPABASE_SERVICE_ROLE_KEY` nos secrets do Streamlit Cloud.")
        st.stop()

    url = st.secrets["SUPABASE_URL"]
    return create_client(url, key)


supabase = init_connection()


# Lista padrão de empresas, usada como fallback quando a tabela no banco ainda não existir.
EMPRESAS_PADRAO = {
    "401": "401 - SST",
    "370": "370 - STI",
    "536": "536 - SCD",
    "570": "570 - SSD",
    "556": "556 - SAC",
}


@st.cache_data(ttl=300)
def carregar_empresas():
    try:
        response = supabase.table("empresas").select("id, nome, ativo").eq("ativo", True).execute()
        dados = response.data or []
        if dados:
            return {str(item["id"]): str(item["nome"]) for item in dados}
    except Exception:
        pass
    return EMPRESAS_PADRAO


EMPRESAS = carregar_empresas()
ARQUIVO_SEED_PARAMETRIZACAO_NOTAS = Path(__file__).with_name("parametrizacao_notas_seed_401.csv")
TABELA_PARAMETRIZACAO_NOTAS = "parametrizacao_notas_empresa"


# --- 2. FUNÇÕES ÚTEIS ---
def tratar_valor(val):
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return abs(float(val))
    val = str(val).replace("R$", "").replace(".", "").replace(",", ".").strip()
    try:
        return abs(float(val))
    except Exception:
        return 0.0


def normalizar_texto(val):
    if pd.isna(val):
        return ""
    return str(val).strip()


def normalizar_chave(val):
    texto = normalizar_texto(val)
    if texto.endswith(".0"):
        texto = texto[:-2]
    return texto


def extrair_codigo_inicial(val):
    texto = normalizar_texto(val)
    if not texto:
        return ""
    match = re.match(r"^\s*(\d+)", texto)
    return match.group(1) if match else ""


def normalizar_chave_relacionamento(val):
    texto = normalizar_texto(val).lower()
    if not texto:
        return ""
    texto = re.sub(r"\s+", " ", texto)
    texto = re.sub(r"\s*-\s*", " - ", texto)
    return texto.strip()


def normalizar_nome_coluna(val):
    texto = normalizar_texto(val)
    if not texto:
        return ""
    texto = re.sub(r"<[^>]+>", " ", texto)
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    texto = texto.lower()
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def aplicar_filtros_cliente(df, col_plano_01, col_nivel4):
    if df is None or df.empty:
        return df

    df_filtrado = df.copy()
    if col_plano_01 and col_plano_01 in df_filtrado.columns:
        valores_plano_01 = df_filtrado[col_plano_01].apply(normalizar_texto).str.upper()
        df_filtrado = df_filtrado[valores_plano_01 != "1 -RECEITAS"]

    if col_nivel4 and col_nivel4 in df_filtrado.columns:
        excluidos_nivel4 = {
            normalizar_chave_relacionamento("2.07.019.003 - TAXA BANCARIA BOLETOS"),
            normalizar_chave_relacionamento("3.02.002.001 - JUROS DE MORA"),
        }
        nivel4_norm = df_filtrado[col_nivel4].apply(normalizar_chave_relacionamento)
        df_filtrado = df_filtrado[~nivel4_norm.isin(excluidos_nivel4)]

    return df_filtrado.copy()


def aplicar_filtros_contabil(df, col_grupo_contabil, col_tipo_lan):
    if df is None or df.empty:
        return df

    df_filtrado = df.copy()
    if col_tipo_lan and col_tipo_lan in df_filtrado.columns:
        tipos = df_filtrado[col_tipo_lan].apply(normalizar_texto).str.upper()
        df_filtrado = df_filtrado[tipos == "D"]

    if col_grupo_contabil and col_grupo_contabil in df_filtrado.columns:
        grupos_excluidos = {
            "TARIFAS BANCARIAS/COBRANCAS",
            "JUROS SOBRE EMPRÉSTIMOS E FINANCIAMENTOS",
            "DEPRECIAÇÕES E AMORTIZAÇÕES",
        }
        grupos = df_filtrado[col_grupo_contabil].apply(normalizar_texto).str.upper()
        df_filtrado = df_filtrado[~grupos.isin(grupos_excluidos)]

    return df_filtrado.copy()


def limpar_colunas(df):
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()
    return df


def obter_coluna(df, candidatos):
    for coluna in candidatos:
        if coluna in df.columns:
            return coluna
    return None


def obter_coluna_flexivel(df, candidatos):
    mapa = {normalizar_nome_coluna(coluna): coluna for coluna in df.columns}
    for candidato in candidatos:
        chave = normalizar_nome_coluna(candidato)
        if chave in mapa:
            return mapa[chave]
    for candidato in candidatos:
        chave = normalizar_nome_coluna(candidato)
        for coluna_norm, coluna_original in mapa.items():
            if chave and (chave in coluna_norm or coluna_norm in chave):
                return coluna_original
    return None


def erro_tabela_ausente(err):
    if not err:
        return False
    if isinstance(err, dict):
        codigo = str(err.get("code", "")).upper()
        mensagem = str(err.get("message", "")).lower()
        return codigo == "PGRST205" or "could not find the table" in mensagem
    texto = str(err).lower()
    return "pgrst205" in texto or "could not find the table" in texto


def erro_rls_policy(err):
    if not err:
        return False
    if isinstance(err, dict):
        codigo = str(err.get("code", "")).upper()
        mensagem = str(err.get("message", "")).lower()
        return codigo == "42501" or "row-level security policy" in mensagem
    texto = str(err).lower()
    return "42501" in texto or "row-level security policy" in texto


def sql_tabela_parametrizacao_notas():
    return f"""create table if not exists public.{TABELA_PARAMETRIZACAO_NOTAS} (
    id bigserial primary key,
    empresa_id text not null references public.empresas(id) on delete cascade,
    plano_conta_nivel_03 text,
    codigo_plano_04 text not null,
    conta_contabil text not null,
    created_at timestamptz not null default now()
);

create index if not exists idx_param_notas_empresa_conta
    on public.{TABELA_PARAMETRIZACAO_NOTAS} (empresa_id, conta_contabil);

create unique index if not exists idx_param_notas_empresa_codigo04_unique
    on public.{TABELA_PARAMETRIZACAO_NOTAS} (empresa_id, codigo_plano_04);"""


def colunas_existentes(df, colunas):
    return [coluna for coluna in colunas if coluna in df.columns]


def formatar_moeda(valor):
    return f"R$ {abs(float(valor)):,.2f}"


def classificar_diferenca(diff):
    if abs(diff) <= 0.01:
        return "🟢 Bateu", "Os valores fecham certinho entre contábil e cliente."
    if diff > 0:
        return "🔴 Divergente", f"Faltam {formatar_moeda(diff)} no cliente para bater com o contábil."
    return "🔴 Divergente", f"Sobram {formatar_moeda(diff)} no cliente em relação ao contábil."


def status_linha(valor, total_ok):
    if pd.isna(valor) or abs(float(valor)) <= 0.0:
        return "Não encontrado"
    return "Encontrado" if total_ok else "Revisar"


def preparar_detalhe_contabil(df_cont_detalhe, col_conta_contabil, col_grupo_contabil, col_historico_contabil, total_ok):
    if df_cont_detalhe.empty:
        return pd.DataFrame(columns=["GRUPO DRE", "HISTORICO", "CONTA DEB", "VALOR", "STATUS"])

    df = df_cont_detalhe.copy()
    df["GRUPO DRE"] = df["_Grupo_Contabil"].apply(normalizar_texto)
    if col_historico_contabil:
        df["HISTORICO"] = df[col_historico_contabil].apply(normalizar_texto)
    else:
        df["HISTORICO"] = ""
    df["CONTA DEB"] = df[col_conta_contabil].apply(normalizar_chave)
    df["VALOR"] = df["Valor_Tratado"].apply(float)
    df["STATUS"] = "Encontrado" if total_ok else "Revisar"

    df = (
        df.groupby(["GRUPO DRE", "HISTORICO", "CONTA DEB"], dropna=False, as_index=False)
        .agg({"VALOR": "sum", "STATUS": "first"})
        .sort_values(["VALOR", "GRUPO DRE", "HISTORICO"], ascending=[False, True, True])
    )
    return df


def preparar_detalhe_cliente(df_cli_detalhe, col_fornecedor_cliente, total_ok):
    if df_cli_detalhe.empty:
        return pd.DataFrame(columns=["CÓD. PLANO 04", "FORNECEDOR", "VALOR", "STATUS"])

    df = df_cli_detalhe.copy()
    df["CÓD. PLANO 04"] = df["_Nivel_4"].apply(normalizar_texto)
    if col_fornecedor_cliente and col_fornecedor_cliente in df.columns:
        df["FORNECEDOR"] = df[col_fornecedor_cliente].apply(normalizar_texto)
    else:
        df["FORNECEDOR"] = df["_Nivel_4"].apply(normalizar_texto)
    df["VALOR"] = df["Valor_Tratado"].apply(float)
    df["STATUS"] = "Encontrado" if total_ok else "Revisar"

    df = (
        df.groupby(["CÓD. PLANO 04", "FORNECEDOR"], dropna=False, as_index=False)
        .agg({"VALOR": "sum", "STATUS": "first"})
        .sort_values(["VALOR", "CÓD. PLANO 04", "FORNECEDOR"], ascending=[False, True, True])
    )
    return df


def montar_quadro_lado_a_lado(df_cont, df_cli):
    max_len = max(len(df_cont), len(df_cli), 1)
    df_cont = df_cont.reset_index(drop=True).reindex(range(max_len))
    df_cli = df_cli.reset_index(drop=True).reindex(range(max_len))

    quadro_export = pd.DataFrame(
        {
            "CONTABIL - GRUPO DRE": df_cont["GRUPO DRE"] if "GRUPO DRE" in df_cont.columns else [""] * max_len,
            "CONTABIL - HISTORICO": df_cont["HISTORICO"] if "HISTORICO" in df_cont.columns else [""] * max_len,
            "CONTABIL - CONTA DEB": df_cont["CONTA DEB"] if "CONTA DEB" in df_cont.columns else [""] * max_len,
            "CONTABIL - VALOR": df_cont["VALOR"] if "VALOR" in df_cont.columns else [0.0] * max_len,
            "CONTABIL - STATUS": df_cont["STATUS"] if "STATUS" in df_cont.columns else [""] * max_len,
            "CLIENTE - CÓD. PLANO 04": df_cli["CÓD. PLANO 04"] if "CÓD. PLANO 04" in df_cli.columns else [""] * max_len,
            "CLIENTE - FORNECEDOR": df_cli["FORNECEDOR"] if "FORNECEDOR" in df_cli.columns else [""] * max_len,
            "CLIENTE - VALOR": df_cli["VALOR"] if "VALOR" in df_cli.columns else [0.0] * max_len,
            "CLIENTE - STATUS": df_cli["STATUS"] if "STATUS" in df_cli.columns else [""] * max_len,
        }
    )

    quadro = quadro_export.copy()
    quadro.columns = pd.MultiIndex.from_tuples([tuple(col.split(" - ", 1)) for col in quadro.columns])
    return quadro, quadro_export


def carregar_parametrizacao_empresa(empresa_id):
    response = (
        supabase.table("parametrizacao_contas")
        .select("id, conta_contabil, fornecedor_cliente")
        .eq("empresa_id", str(empresa_id))
        .execute()
    )
    df_banco = pd.DataFrame(response.data)
    if df_banco.empty:
        df_banco = pd.DataFrame(columns=["id", "conta_contabil", "fornecedor_cliente"])
    return df_banco


def carregar_parametrizacao_notas_empresa(empresa_id):
    try:
        response = (
            supabase.table(TABELA_PARAMETRIZACAO_NOTAS)
            .select("id, plano_conta_nivel_03, codigo_plano_04, conta_contabil")
            .eq("empresa_id", str(empresa_id))
            .execute()
        )
        df_banco = pd.DataFrame(response.data)
        st.session_state["erro_parametrizacao_notas"] = None
    except Exception as err:
        st.session_state["erro_parametrizacao_notas"] = err
        df_banco = pd.DataFrame()

    if df_banco.empty:
        df_banco = pd.DataFrame(columns=["id", "plano_conta_nivel_03", "codigo_plano_04", "conta_contabil"])
    return df_banco


def carregar_seed_parametrizacao_notas():
    if not ARQUIVO_SEED_PARAMETRIZACAO_NOTAS.exists():
        return pd.DataFrame(columns=["empresa_id", "plano_conta_nivel_03", "codigo_plano_04", "conta_contabil"])
    df_seed = pd.read_csv(ARQUIVO_SEED_PARAMETRIZACAO_NOTAS, dtype=str).fillna("")
    for coluna in ["empresa_id", "plano_conta_nivel_03", "codigo_plano_04", "conta_contabil"]:
        if coluna not in df_seed.columns:
            df_seed[coluna] = ""
    return df_seed[["empresa_id", "plano_conta_nivel_03", "codigo_plano_04", "conta_contabil"]].copy()


def preparar_pendencias_parametrizacao(
    df_cliente,
    df_parametros,
    col_nivel3_cliente,
    col_conta_cliente,
    col_fornecedor_cliente=None,
    col_descricao_cliente=None,
):
    if df_cliente is None or df_cliente.empty:
        return pd.DataFrame(columns=["GRUPO DRE", "CÓD. PLANO 04", "FORNECEDOR", "VALOR", "QTD LANCAMENTOS", "STATUS"])

    df = df_cliente.copy()
    if "_Nivel_3" not in df.columns and col_nivel3_cliente in df.columns:
        df["_Nivel_3"] = df[col_nivel3_cliente].apply(normalizar_texto)
    if "_Nivel_4" not in df.columns and col_conta_cliente in df.columns:
        df["_Nivel_4"] = df[col_conta_cliente].apply(normalizar_texto)
    if "_Fornecedor" not in df.columns and col_fornecedor_cliente and col_fornecedor_cliente in df.columns:
        df["_Fornecedor"] = df[col_fornecedor_cliente].apply(normalizar_texto)
    if col_descricao_cliente and col_descricao_cliente in df.columns and "_Descricao" not in df.columns:
        df["_Descricao"] = df[col_descricao_cliente].apply(normalizar_texto)

    parametros_existentes = set()
    if df_parametros is not None and not df_parametros.empty and "fornecedor_cliente" in df_parametros.columns:
        parametros_existentes = set(
            df_parametros["fornecedor_cliente"].dropna().astype(str).map(normalizar_chave_relacionamento)
        )

    df["__parametrizado"] = df["_Nivel_4"].apply(normalizar_chave_relacionamento).isin(parametros_existentes)
    df_pend = df[~df["__parametrizado"]].copy()

    if df_pend.empty:
        return pd.DataFrame(columns=["GRUPO DRE", "CÓD. PLANO 04", "FORNECEDOR", "VALOR", "QTD LANCAMENTOS", "STATUS"])

    df_pend["FORNECEDOR"] = df_pend["_Fornecedor"] if "_Fornecedor" in df_pend.columns else df_pend["_Nivel_4"]
    df_pend["VALOR"] = df_pend["Valor_Tratado"].apply(float)
    df_pend["STATUS"] = "Sem parametrização"

    df_pend = (
        df_pend.groupby(["_Nivel_3", "_Nivel_4", "FORNECEDOR"], dropna=False, as_index=False)
        .agg(QTD_LANCAMENTOS=("FORNECEDOR", "size"), VALOR=("VALOR", "sum"))
        .rename(columns={"_Nivel_3": "GRUPO DRE", "_Nivel_4": "CÓD. PLANO 04"})
        .sort_values(["VALOR", "QTD_LANCAMENTOS"], ascending=False)
    )

    df_pend["STATUS"] = "Sem parametrização"
    return df_pend


def preparar_base_documentos_fiscais(df_nfe, df_nfse):
    bases = []
    avisos = []

    if df_nfse is not None and not df_nfse.empty:
        col_plano_nfse = obter_coluna_flexivel(df_nfse, ["Plano de contas", "Plano de Contas"])
        col_valor_nfse = obter_coluna_flexivel(df_nfse, ["Vr. Total"])
        col_fornecedor_nfse = obter_coluna_flexivel(df_nfse, ["Fornecedor"])
        col_numero_nfse = obter_coluna_flexivel(df_nfse, ["Nr. Nota"])

        faltantes_nfse = []
        if not col_plano_nfse:
            faltantes_nfse.append("Plano de contas")
        if not col_valor_nfse:
            faltantes_nfse.append("Vr. Total")
        if not col_fornecedor_nfse:
            faltantes_nfse.append("Fornecedor")

        if faltantes_nfse:
            avisos.append("NFSe sem colunas esperadas: " + ", ".join(faltantes_nfse))
        else:
            base_nfse = df_nfse.copy()
            base_nfse["_Nivel_4"] = base_nfse[col_plano_nfse].apply(normalizar_texto)
            base_nfse["_Nivel_4_Norm"] = base_nfse["_Nivel_4"].apply(normalizar_chave_relacionamento)
            base_nfse["_Fornecedor"] = base_nfse[col_fornecedor_nfse].apply(normalizar_texto)
            base_nfse["_Documento"] = (
                base_nfse[col_numero_nfse].apply(normalizar_texto) if col_numero_nfse else ""
            )
            base_nfse["Valor_Tratado"] = base_nfse[col_valor_nfse].apply(tratar_valor)
            base_nfse["_Origem"] = "NFSe"
            bases.append(base_nfse)

    if df_nfe is not None and not df_nfe.empty:
        col_plano_nfe = obter_coluna_flexivel(df_nfe, ["Plano de Contas", "Plano de contas", "TAG de Cliente"])
        col_valor_nfe = obter_coluna_flexivel(df_nfe, ["Vr. Nota"])
        col_fornecedor_nfe = obter_coluna_flexivel(df_nfe, ["Razão social (Emitente)", "Razo social (Emitente)"])
        col_numero_nfe = obter_coluna_flexivel(df_nfe, ["Nr. NFe", "Documento"])

        faltantes_nfe = []
        if not col_plano_nfe:
            faltantes_nfe.append("Plano de Contas")
        if not col_valor_nfe:
            faltantes_nfe.append("Vr. Nota")
        if not col_fornecedor_nfe:
            faltantes_nfe.append("Razão social (Emitente)")

        if faltantes_nfe:
            avisos.append("NFe sem colunas esperadas: " + ", ".join(faltantes_nfe))
        else:
            base_nfe = df_nfe.copy()
            base_nfe["_Nivel_4"] = base_nfe[col_plano_nfe].apply(normalizar_texto)
            validos_nivel4 = base_nfe["_Nivel_4"].str.contains(r"^\d+\.\d+\.\d+\.\d+\s+-\s+", regex=True, na=False)
            if validos_nivel4.sum() == 0:
                avisos.append("NFe sem coluna útil de Plano de Contas; arquivo ignorado nesta análise.")
            else:
                base_nfe = base_nfe[validos_nivel4].copy()
                base_nfe["_Nivel_4"] = base_nfe["_Nivel_4"].apply(normalizar_texto)
                base_nfe["_Nivel_4_Norm"] = base_nfe["_Nivel_4"].apply(normalizar_chave_relacionamento)
                base_nfe["_Fornecedor"] = base_nfe[col_fornecedor_nfe].apply(normalizar_texto)
                base_nfe["_Documento"] = base_nfe[col_numero_nfe].apply(normalizar_texto) if col_numero_nfe else ""
                base_nfe["Valor_Tratado"] = base_nfe[col_valor_nfe].apply(tratar_valor)
                base_nfe["_Origem"] = "NFe"
                bases.append(base_nfe)

    if not bases:
        return pd.DataFrame(), avisos

    df_docs = pd.concat(bases, ignore_index=True)
    df_docs = df_docs[df_docs["_Nivel_4"].astype(str).str.strip().ne("")].copy()
    return df_docs, avisos


def calcular_totais_relatorios_notas(arquivo_contabil, arquivo_nfse, arquivo_nfe):
    totais = {
        "contabil": 0.0,
        "nfse": 0.0,
        "nfe": 0.0,
    }
    avisos = []

    if arquivo_contabil:
        arquivo_contabil.seek(0)
        df_contabil = limpar_colunas(pd.read_excel(arquivo_contabil, engine="openpyxl"))
        col_valor_contabil = obter_coluna_flexivel(df_contabil, ["valdeb"])
        col_tipo_lan_contabil = obter_coluna_flexivel(df_contabil, ["tipo_lan"])

        if not col_valor_contabil:
            avisos.append("Razão contábil sem coluna `valdeb` para montar o quadro geral.")
        else:
            if col_tipo_lan_contabil:
                df_contabil = df_contabil[
                    df_contabil[col_tipo_lan_contabil].apply(normalizar_texto).str.upper() == "D"
                ].copy()
            totais["contabil"] = round(df_contabil[col_valor_contabil].apply(tratar_valor).sum(), 2)
        arquivo_contabil.seek(0)

    if arquivo_nfse:
        arquivo_nfse.seek(0)
        df_nfse = limpar_colunas(pd.read_excel(arquivo_nfse, engine="openpyxl"))
        col_valor_nfse = obter_coluna_flexivel(df_nfse, ["Vr. Total"])
        if not col_valor_nfse:
            avisos.append("NFSe sem coluna `Vr. Total` para montar o quadro geral.")
        else:
            totais["nfse"] = round(df_nfse[col_valor_nfse].apply(tratar_valor).sum(), 2)
        arquivo_nfse.seek(0)

    if arquivo_nfe:
        arquivo_nfe.seek(0)
        df_nfe = limpar_colunas(pd.read_excel(arquivo_nfe, engine="openpyxl"))
        col_valor_nfe = obter_coluna_flexivel(df_nfe, ["Vr. Nota"])
        if not col_valor_nfe:
            avisos.append("NFe sem coluna `Vr. Nota` para montar o quadro geral.")
        else:
            totais["nfe"] = round(df_nfe[col_valor_nfe].apply(tratar_valor).sum(), 2)
        arquivo_nfe.seek(0)

    totais["documentos"] = round(totais["nfse"] + totais["nfe"], 2)
    return totais, avisos


def preparar_detalhe_documentos(df_docs_detalhe, total_ok):
    if df_docs_detalhe.empty:
        return pd.DataFrame(columns=["PLANO DE CONTAS", "FORNECEDOR", "ORIGEM", "DOCUMENTO", "VALOR", "STATUS"])

    df = df_docs_detalhe.copy()
    df["PLANO DE CONTAS"] = df["_Nivel_4"].apply(normalizar_texto)
    df["FORNECEDOR"] = df["_Fornecedor"].apply(normalizar_texto)
    df["ORIGEM"] = df["_Origem"].apply(normalizar_texto)
    df["DOCUMENTO"] = df["_Documento"].apply(normalizar_texto)
    df["VALOR"] = df["Valor_Tratado"].apply(float)
    df["STATUS"] = "Encontrado" if total_ok else "Revisar"

    return (
        df.groupby(["PLANO DE CONTAS", "FORNECEDOR", "ORIGEM", "DOCUMENTO"], dropna=False, as_index=False)
        .agg({"VALOR": "sum", "STATUS": "first"})
        .sort_values(["VALOR", "PLANO DE CONTAS", "FORNECEDOR"], ascending=[False, True, True])
    )


def montar_quadro_lado_a_lado_documentos(df_cont, df_docs):
    max_len = max(len(df_cont), len(df_docs), 1)
    df_cont = df_cont.reset_index(drop=True).reindex(range(max_len))
    df_docs = df_docs.reset_index(drop=True).reindex(range(max_len))

    quadro_export = pd.DataFrame(
        {
            "CONTABIL - GRUPO DRE": df_cont["GRUPO DRE"] if "GRUPO DRE" in df_cont.columns else [""] * max_len,
            "CONTABIL - HISTORICO": df_cont["HISTORICO"] if "HISTORICO" in df_cont.columns else [""] * max_len,
            "CONTABIL - CONTA DEB": df_cont["CONTA DEB"] if "CONTA DEB" in df_cont.columns else [""] * max_len,
            "CONTABIL - VALOR": df_cont["VALOR"] if "VALOR" in df_cont.columns else [0.0] * max_len,
            "CONTABIL - STATUS": df_cont["STATUS"] if "STATUS" in df_cont.columns else [""] * max_len,
            "DOCUMENTOS - PLANO DE CONTAS": df_docs["PLANO DE CONTAS"] if "PLANO DE CONTAS" in df_docs.columns else [""] * max_len,
            "DOCUMENTOS - FORNECEDOR": df_docs["FORNECEDOR"] if "FORNECEDOR" in df_docs.columns else [""] * max_len,
            "DOCUMENTOS - ORIGEM": df_docs["ORIGEM"] if "ORIGEM" in df_docs.columns else [""] * max_len,
            "DOCUMENTOS - DOCUMENTO": df_docs["DOCUMENTO"] if "DOCUMENTO" in df_docs.columns else [""] * max_len,
            "DOCUMENTOS - VALOR": df_docs["VALOR"] if "VALOR" in df_docs.columns else [0.0] * max_len,
            "DOCUMENTOS - STATUS": df_docs["STATUS"] if "STATUS" in df_docs.columns else [""] * max_len,
        }
    )

    quadro = quadro_export.copy()
    quadro.columns = pd.MultiIndex.from_tuples([tuple(col.split(" - ", 1)) for col in quadro.columns])
    return quadro, quadro_export


def preparar_pendencias_documentos(df_docs, df_parametros):
    if df_docs is None or df_docs.empty:
        return pd.DataFrame(
            columns=["PLANO DE CONTA Nº. 03", "CÓD. PLANO DE CONTA Nº. 04", "FORNECEDOR/EMITENTE", "ORIGENS"]
        )

    df = df_docs.copy()
    parametros_existentes = set()
    if df_parametros is not None and not df_parametros.empty and "codigo_plano_04" in df_parametros.columns:
        parametros_existentes = set(
            df_parametros["codigo_plano_04"].dropna().astype(str).map(normalizar_chave_relacionamento)
        )

    df["__parametrizado"] = df["_Nivel_4_Norm"].isin(parametros_existentes)
    df_pend = df[~df["__parametrizado"]].copy()
    if df_pend.empty:
        return pd.DataFrame(
            columns=["PLANO DE CONTA Nº. 03", "CÓD. PLANO DE CONTA Nº. 04", "FORNECEDOR/EMITENTE", "ORIGENS"]
        )

    juntar_unicos = lambda serie: " | ".join(
        [item for item in pd.Series(serie).dropna().astype(str).map(normalizar_texto).unique() if item]
    )

    df_pend = df_pend.drop_duplicates(subset=["_Nivel_4", "_Fornecedor", "_Origem", "_Documento"])
    df_pend = (
        df_pend.groupby("_Nivel_4", dropna=False, as_index=False)
        .agg(
            **{
                "FORNECEDOR/EMITENTE": ("_Fornecedor", juntar_unicos),
                "ORIGENS": ("_Origem", juntar_unicos),
            }
        )
        .rename(columns={"_Nivel_4": "CÓD. PLANO DE CONTA Nº. 04"})
        .sort_values(["CÓD. PLANO DE CONTA Nº. 04"], ascending=True)
    )
    df_pend.insert(0, "PLANO DE CONTA Nº. 03", "")
    return df_pend


def exportar_modelo_parametrizacao_notas(nome_empresa, df_modelo):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export = df_modelo.copy()
        df_export.to_excel(writer, sheet_name="Modelo", index=False, startrow=2)

        wb = writer.book
        ws = wb["Modelo"]

        azul = PatternFill("solid", fgColor="DCE6F1")
        cinza = PatternFill("solid", fgColor="F3F6F9")
        titulo = PatternFill("solid", fgColor="2F5597")
        branco = Font(color="FFFFFF", bold=True)
        bold = Font(bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df_export.columns), 1))
        ws["A1"] = f"MODELO DE PARAMETRIZACAO - {nome_empresa}"
        ws["A1"].fill = titulo
        ws["A1"].font = branco
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws["A2"] = "Preencha apenas a coluna CONTA CONTÁBIL e reimporte a planilha no sistema."
        ws["A2"].font = Font(italic=True)

        ws.freeze_panes = "A4"
        for cell in ws[3]:
            cell.font = bold
            cell.fill = azul
            cell.border = border

        for row in ws.iter_rows(min_row=4):
            for cell in row:
                cell.border = border
                if cell.column != 1:
                    cell.fill = cinza

        for col in ws.columns:
            largura = 0
            letra = get_column_letter(col[0].column)
            for cell in col:
                try:
                    largura = max(largura, len(str(cell.value)) if cell.value is not None else 0)
                except Exception:
                    pass
            ws.column_dimensions[letra].width = min(max(largura + 2, 18), 42)

    return output.getvalue()


def preparar_registros_parametrizacao_notas(df_parametrizacao, empresa_id):
    novos_dados = []
    vistos = set()
    if df_parametrizacao is None or df_parametrizacao.empty:
        return novos_dados

    for _, row in df_parametrizacao.iterrows():
        plano_03 = normalizar_texto(row.get("PLANO DE CONTA Nº. 03", ""))
        codigo_04 = normalizar_texto(row.get("CÓD. PLANO DE CONTA Nº. 04", ""))
        conta = normalizar_chave(row.get("CONTA CONTÁBIL", ""))
        if not codigo_04 or not conta:
            continue
        chave = (str(empresa_id), codigo_04)
        if chave in vistos:
            continue
        vistos.add(chave)
        novos_dados.append(
            {
                "empresa_id": str(empresa_id),
                "plano_conta_nivel_03": plano_03,
                "codigo_plano_04": codigo_04,
                "conta_contabil": conta,
            }
        )
    return novos_dados


def importar_modelo_parametrizacao_notas(arquivo):
    candidatos_conta = ["CONTA CONTÁBIL", "CONTA CONTABIL"]
    candidatos_plano_03 = ["PLANO DE CONTA Nº. 03", "PLANO DE CONTA N 03", "PLANO DE CONTA Nº. 03"]
    candidatos_codigo_04 = [
        "CÓD. PLANO DE CONTA Nº. 04",
        "COD. PLANO DE CONTA Nº. 04",
        "COD. PLANO DE CONTA N 04",
        "CÓD. PLANO 04",
    ]

    def normalizar_dataframe_importacao(df_importado):
        df_importado = limpar_colunas(df_importado)
        colunas_mapeadas = {
            "CONTA CONTÁBIL": obter_coluna_flexivel(df_importado, candidatos_conta),
            "PLANO DE CONTA Nº. 03": obter_coluna_flexivel(df_importado, candidatos_plano_03),
            "CÓD. PLANO DE CONTA Nº. 04": obter_coluna_flexivel(df_importado, candidatos_codigo_04),
        }
        return df_importado, colunas_mapeadas

    arquivo.seek(0)
    bruto = pd.read_excel(arquivo, engine="openpyxl", header=None)

    melhor_tentativa = None
    for idx in range(min(len(bruto), 8)):
        cabecalho = [normalizar_texto(valor) for valor in bruto.iloc[idx].tolist()]
        if not any(cabecalho):
            continue

        df_candidato = bruto.iloc[idx + 1 :].copy()
        df_candidato.columns = cabecalho
        df_candidato = df_candidato.dropna(how="all")
        df_candidato, colunas_mapeadas = normalizar_dataframe_importacao(df_candidato)

        if colunas_mapeadas.get("CONTA CONTÁBIL") and colunas_mapeadas.get("CÓD. PLANO DE CONTA Nº. 04"):
            melhor_tentativa = (df_candidato, colunas_mapeadas)
            break

    if melhor_tentativa is None:
        arquivo.seek(0)
        df_padrao = pd.read_excel(arquivo, engine="openpyxl")
        df_padrao, colunas_mapeadas = normalizar_dataframe_importacao(df_padrao)
        melhor_tentativa = (df_padrao, colunas_mapeadas)

    df_importado, colunas_mapeadas = melhor_tentativa

    obrigatorias = ["CONTA CONTÁBIL", "CÓD. PLANO DE CONTA Nº. 04"]
    faltando = [coluna for coluna in obrigatorias if not colunas_mapeadas.get(coluna)]
    if faltando:
        raise ValueError(
            "A planilha importada não contém as colunas obrigatórias: "
            f"{', '.join(faltando)}. Use o modelo baixado no sistema."
        )

    df_normalizado = pd.DataFrame(
        {
            "CONTA CONTÁBIL": df_importado[colunas_mapeadas["CONTA CONTÁBIL"]],
            "PLANO DE CONTA Nº. 03": (
                df_importado[colunas_mapeadas["PLANO DE CONTA Nº. 03"]]
                if colunas_mapeadas.get("PLANO DE CONTA Nº. 03")
                else ""
            ),
            "CÓD. PLANO DE CONTA Nº. 04": df_importado[colunas_mapeadas["CÓD. PLANO DE CONTA Nº. 04"]],
        }
    )
    return df_normalizado.dropna(how="all").fillna("")


def salvar_parametrizacao_notas(df_parametrizacao, empresa_id):
    novos_dados = preparar_registros_parametrizacao_notas(df_parametrizacao, empresa_id)
    if novos_dados:
        supabase.table(TABELA_PARAMETRIZACAO_NOTAS).upsert(
            novos_dados,
            on_conflict="empresa_id,codigo_plano_04",
        ).execute()
    return len(novos_dados)


def exportar_excel_relatorio(nome_cliente, df_resumo, quadro_export):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        resumo_export = df_resumo.copy()
        resumo_export.to_excel(writer, sheet_name="Resumo", index=False)

        quadro_export.to_excel(writer, sheet_name="Detalhe", index=False, startrow=2)

        wb = writer.book

        # Estilos leves para leitura.
        azul = PatternFill("solid", fgColor="DCE6F1")
        verde = PatternFill("solid", fgColor="E2F0D9")
        cinza = PatternFill("solid", fgColor="F3F6F9")
        titulo = PatternFill("solid", fgColor="2F5597")
        branco = Font(color="FFFFFF", bold=True)
        bold = Font(bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws = wb["Resumo"]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = bold
            cell.fill = cinza
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border

        for col in ws.columns:
            largura = 0
            letra = get_column_letter(col[0].column)
            for cell in col:
                try:
                    largura = max(largura, len(str(cell.value)) if cell.value is not None else 0)
                except Exception:
                    pass
            ws.column_dimensions[letra].width = min(max(largura + 2, 12), 28)

        wd = wb["Detalhe"]
        wd.freeze_panes = "A4"
        wd.merge_cells("A1:I1")
        wd["A1"] = f"NOME DO CLIENTE: {nome_cliente}"
        wd["A1"].fill = titulo
        wd["A1"].font = branco
        wd["A1"].alignment = Alignment(horizontal="center", vertical="center")

        wd.merge_cells("A2:E2")
        wd.merge_cells("F2:I2")
        wd["A2"] = "CONTABIL"
        wd["F2"] = "CLIENTE"
        wd["A2"].fill = azul
        wd["F2"].fill = verde
        wd["A2"].font = bold
        wd["F2"].font = bold
        wd["A2"].alignment = Alignment(horizontal="center")
        wd["F2"].alignment = Alignment(horizontal="center")

        for cell in wd[3]:
            cell.font = bold
            cell.fill = cinza
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in wd.iter_rows(min_row=4):
            for cell in row:
                cell.border = border
                if cell.column in [2, 3, 4, 5]:
                    cell.fill = azul
                elif cell.column in [6, 7, 8, 9]:
                    cell.fill = verde

        for row in wd.iter_rows(min_row=4):
            for cell in row:
                if cell.column in [5, 9]:
                    valor = str(cell.value).lower() if cell.value is not None else ""
                    if "bateu" in valor or "encontrado" in valor:
                        cell.fill = PatternFill("solid", fgColor="E2F0D9")
                    elif "diverg" in valor or "revis" in valor:
                        cell.fill = PatternFill("solid", fgColor="FCE4D6")
                    elif "não" in valor or "nao" in valor:
                        cell.fill = PatternFill("solid", fgColor="FFF2CC")

        for col in wd.columns:
            largura = 0
            letra = get_column_letter(col[0].column)
            for cell in col:
                try:
                    largura = max(largura, len(str(cell.value)) if cell.value is not None else 0)
                except Exception:
                    pass
            wd.column_dimensions[letra].width = min(max(largura + 2, 14), 35)

        # Formatação numérica
        for row in wd.iter_rows(min_row=4):
            for cell in row:
                if cell.column in [4, 8]:
                    cell.number_format = 'R$ #,##0.00'

    output.seek(0)
    return output.getvalue()


def exportar_excel_relatorio_documentos(nome_cliente, df_resumo, quadro_export):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
        quadro_export.to_excel(writer, sheet_name="Detalhe", index=False, startrow=2)

        wb = writer.book
        azul = PatternFill("solid", fgColor="DCE6F1")
        verde = PatternFill("solid", fgColor="E2F0D9")
        cinza = PatternFill("solid", fgColor="F3F6F9")
        titulo = PatternFill("solid", fgColor="2F5597")
        branco = Font(color="FFFFFF", bold=True)
        bold = Font(bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws = wb["Resumo"]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = bold
            cell.fill = cinza
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border

        wd = wb["Detalhe"]
        wd.freeze_panes = "A4"
        wd.merge_cells("A1:K1")
        wd["A1"] = f"NOME DO CLIENTE: {nome_cliente}"
        wd["A1"].fill = titulo
        wd["A1"].font = branco
        wd["A1"].alignment = Alignment(horizontal="center", vertical="center")
        wd.merge_cells("A2:E2")
        wd.merge_cells("F2:K2")
        wd["A2"] = "CONTABIL"
        wd["F2"] = "DOCUMENTOS"
        wd["A2"].fill = azul
        wd["F2"].fill = verde
        wd["A2"].font = bold
        wd["F2"].font = bold
        wd["A2"].alignment = Alignment(horizontal="center")
        wd["F2"].alignment = Alignment(horizontal="center")

        for cell in wd[3]:
            cell.font = bold
            cell.fill = cinza
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in wd.iter_rows(min_row=4):
            for cell in row:
                cell.border = border
                if 1 <= cell.column <= 5:
                    cell.fill = azul
                elif 6 <= cell.column <= 11:
                    cell.fill = verde

        for row in wd.iter_rows(min_row=4):
            for cell in row:
                if cell.column in [5, 11]:
                    valor = str(cell.value).lower() if cell.value is not None else ""
                    if "bateu" in valor or "encontrado" in valor:
                        cell.fill = PatternFill("solid", fgColor="E2F0D9")
                    elif "diverg" in valor or "revis" in valor:
                        cell.fill = PatternFill("solid", fgColor="FCE4D6")
                    elif "não" in valor or "nao" in valor:
                        cell.fill = PatternFill("solid", fgColor="FFF2CC")

        for col in wd.columns:
            largura = 0
            letra = get_column_letter(col[0].column)
            for cell in col:
                try:
                    largura = max(largura, len(str(cell.value)) if cell.value is not None else 0)
                except Exception:
                    pass
            wd.column_dimensions[letra].width = min(max(largura + 2, 14), 35)

        for row in wd.iter_rows(min_row=4):
            for cell in row:
                if cell.column in [4, 10]:
                    cell.number_format = 'R$ #,##0.00'

    output.seek(0)
    return output.getvalue()


def assinatura_arquivos(*arquivos):
    return tuple(
        (
            getattr(arquivo, "name", None),
            getattr(arquivo, "size", None),
        )
        if arquivo is not None
        else (None, None)
        for arquivo in arquivos
    )


def renderizar_analise_conferencia(analise):
    df_resultados = analise["df_resultados"].copy()
    detalhes_por_conta = analise["detalhes_por_conta"]
    empresa_id = analise["empresa_id"]
    empresa_nome = analise["empresa_nome"]
    col_conta_contabil = analise["col_conta_contabil"]
    col_grupo_contabil = analise["col_grupo_contabil"]
    col_historico_contabil = analise["col_historico_contabil"]
    col_fornecedor_cliente = analise["col_fornecedor_cliente"]

    if df_resultados.empty:
        st.info("Nenhuma conta com valor diferente de zero foi encontrada para comparar.")
        return

    st.markdown("---")
    tot_batidos = len(df_resultados[df_resultados["STATUS"] == "🟢 Bateu"])
    tot_divergentes = len(df_resultados[df_resultados["STATUS"] == "🔴 Divergente"])
    tot_sem_map = len(df_resultados[df_resultados["STATUS"] == "⚠️ Sem parametrização"])

    c1, c2, c3 = st.columns(3)
    c1.metric("🟢 Bateu", tot_batidos)
    c2.metric("🔴 Divergências", tot_divergentes)
    c3.metric("⚠️ Sem parametrização", tot_sem_map)

    filtro_key = f"filtro_visao_{empresa_id}"
    if filtro_key not in st.session_state:
        st.session_state[filtro_key] = "Todos"

    filtro = st.radio(
        "Filtrar Visão:",
        ["Todos", "🔴 Divergente", "⚠️ Sem parametrização", "🟢 Bateu"],
        horizontal=True,
        key=filtro_key,
    )
    df_filtrado = df_resultados
    if filtro != "Todos":
        df_filtrado = df_filtrado[df_filtrado["STATUS"] == filtro]

    if df_filtrado.empty:
        st.info("Nenhuma conta corresponde ao filtro selecionado.")
        return

    colunas_visiveis = [
        "Conta Débito",
        "Grupo DRE (Contábil)",
        "Valor Contábil",
        "Valor Cliente",
        "Diferença",
        "STATUS",
        "MOTIVO",
    ]
    st.dataframe(
        df_filtrado[colunas_visiveis].style.format(
            {
                "Valor Contábil": "R$ {:,.2f}",
                "Valor Cliente": "R$ {:,.2f}",
                "Diferença": "R$ {:,.2f}",
            }
        ),
        use_container_width=True,
    )

    # --- DETALHAMENTO ANALÍTICO ---
    st.markdown("---")
    st.subheader("Quadro dos Lançamentos")
    st.caption("Contábil e cliente lado a lado, com o total da conta selecionada.")

    contas_visiveis = df_filtrado["Conta Débito"].tolist()
    conta_key = f"conta_detalhe_{empresa_id}"
    if conta_key not in st.session_state or st.session_state[conta_key] not in contas_visiveis:
        st.session_state[conta_key] = contas_visiveis[0]

    conta_detalhe = st.selectbox("Conta para detalhar", contas_visiveis, key=conta_key)

    detalhe_atual = detalhes_por_conta.get(
        conta_detalhe, {"contabil": pd.DataFrame(), "cliente": pd.DataFrame()}
    )
    df_cont_detalhe = detalhe_atual["contabil"]
    df_cli_detalhe = detalhe_atual["cliente"]
    resumo_linha = df_filtrado[df_filtrado["Conta Débito"] == conta_detalhe].iloc[0]
    total_ok = resumo_linha["STATUS"] == "🟢 Bateu"

    st.markdown(f"### NOME DO CLIENTE: {empresa_nome}")
    st.markdown("#### CONTABIL | CLIENTE")

    df_cont_rel = preparar_detalhe_contabil(
        df_cont_detalhe, col_conta_contabil, col_grupo_contabil, col_historico_contabil, total_ok
    )
    df_cli_rel = preparar_detalhe_cliente(df_cli_detalhe, col_fornecedor_cliente, total_ok)
    quadro_display, quadro_export = montar_quadro_lado_a_lado(df_cont_rel, df_cli_rel)

    st.dataframe(
        quadro_display.style.format(
            {
                ("CONTABIL", "VALOR"): "R$ {:,.2f}",
                ("CLIENTE", "VALOR"): "R$ {:,.2f}",
            }
        ),
        use_container_width=True,
        hide_index=True,
    )

    resumo_export = df_resultados[
        [
            "Conta Débito",
            "Grupo DRE (Contábil)",
            "Valor Contábil",
            "Valor Cliente",
            "Diferença",
            "STATUS",
            "MOTIVO",
            "QTD CONTÁBIL",
            "QTD CLIENTE",
            "DIF. QTD",
        ]
    ].copy()

    arquivo_excel = exportar_excel_relatorio(empresa_nome, resumo_export, quadro_export)

    st.download_button(
        "📥 Exportar relatório em Excel",
        data=arquivo_excel,
        file_name=f"conciliacao_{empresa_id}_{conta_detalhe}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


def renderizar_analise_documentos(analise):
    df_resultados = analise["df_resultados"].copy()
    detalhes_por_conta = analise["detalhes_por_conta"]
    empresa_id = analise["empresa_id"]
    empresa_nome = analise["empresa_nome"]
    col_conta_contabil = analise["col_conta_contabil"]
    col_grupo_contabil = analise["col_grupo_contabil"]
    col_historico_contabil = analise["col_historico_contabil"]

    if df_resultados.empty:
        st.info("Nenhuma conta com valor diferente de zero foi encontrada para comparar.")
        return

    st.markdown("---")
    tot_batidos = len(df_resultados[df_resultados["STATUS"] == "🟢 Bateu"])
    tot_divergentes = len(df_resultados[df_resultados["STATUS"] == "🔴 Divergente"])
    tot_sem_map = len(df_resultados[df_resultados["STATUS"] == "⚠️ Sem parametrização"])

    c1, c2, c3 = st.columns(3)
    c1.metric("🟢 Bateu", tot_batidos)
    c2.metric("🔴 Divergências", tot_divergentes)
    c3.metric("⚠️ Sem parametrização", tot_sem_map)

    filtro_key = f"filtro_visao_documentos_{empresa_id}"
    if filtro_key not in st.session_state:
        st.session_state[filtro_key] = "Todos"

    filtro = st.radio(
        "Filtrar Visão:",
        ["Todos", "🔴 Divergente", "⚠️ Sem parametrização", "🟢 Bateu"],
        horizontal=True,
        key=filtro_key,
    )
    df_filtrado = df_resultados
    if filtro != "Todos":
        df_filtrado = df_filtrado[df_filtrado["STATUS"] == filtro]

    if df_filtrado.empty:
        st.info("Nenhuma conta corresponde ao filtro selecionado.")
        return

    st.dataframe(
        df_filtrado[
            [
                "Conta Débito",
                "Grupo DRE (Contábil)",
                "Valor Contábil",
                "Valor Documentos",
                "Diferença",
                "STATUS",
                "MOTIVO",
            ]
        ].style.format(
            {
                "Valor Contábil": "R$ {:,.2f}",
                "Valor Documentos": "R$ {:,.2f}",
                "Diferença": "R$ {:,.2f}",
            }
        ),
        use_container_width=True,
    )

    st.markdown("---")
    st.subheader("Quadro dos Lançamentos")
    st.caption("Contábil e documentos fiscais lado a lado, com o total da conta selecionada.")

    contas_visiveis = df_filtrado["Conta Débito"].tolist()
    conta_key = f"conta_detalhe_documentos_{empresa_id}"
    if conta_key not in st.session_state or st.session_state[conta_key] not in contas_visiveis:
        st.session_state[conta_key] = contas_visiveis[0]

    conta_detalhe = st.selectbox("Conta para detalhar", contas_visiveis, key=conta_key)
    detalhe_atual = detalhes_por_conta.get(conta_detalhe, {"contabil": pd.DataFrame(), "documentos": pd.DataFrame()})
    df_cont_detalhe = detalhe_atual["contabil"]
    df_docs_detalhe = detalhe_atual["documentos"]
    resumo_linha = df_filtrado[df_filtrado["Conta Débito"] == conta_detalhe].iloc[0]
    total_ok = resumo_linha["STATUS"] == "🟢 Bateu"

    st.markdown(f"### NOME DO CLIENTE: {empresa_nome}")
    st.markdown("#### CONTABIL | DOCUMENTOS")

    df_cont_rel = preparar_detalhe_contabil(
        df_cont_detalhe, col_conta_contabil, col_grupo_contabil, col_historico_contabil, total_ok
    )
    df_docs_rel = preparar_detalhe_documentos(df_docs_detalhe, total_ok)
    quadro_display, quadro_export = montar_quadro_lado_a_lado_documentos(df_cont_rel, df_docs_rel)

    st.dataframe(
        quadro_display.style.format(
            {
                ("CONTABIL", "VALOR"): "R$ {:,.2f}",
                ("DOCUMENTOS", "VALOR"): "R$ {:,.2f}",
            }
        ),
        use_container_width=True,
        hide_index=True,
    )

    resumo_export = df_resultados[
        [
            "Conta Débito",
            "Grupo DRE (Contábil)",
            "Valor Contábil",
            "Valor Documentos",
            "Diferença",
            "STATUS",
            "MOTIVO",
            "QTD CONTÁBIL",
            "QTD DOCUMENTOS",
        ]
    ].copy()
    arquivo_excel = exportar_excel_relatorio_documentos(empresa_nome, resumo_export, quadro_export)
    st.download_button(
        "📥 Exportar relatório em Excel",
        data=arquivo_excel,
        file_name=f"conciliacao_documentos_{empresa_id}_{conta_detalhe}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


# --- 3. MENU LATERAL ---
st.sidebar.title("Navegação")
empresa_selecionada = st.sidebar.selectbox(
    "🏢 Selecione a Empresa:",
    list(EMPRESAS.keys()),
    format_func=lambda x: EMPRESAS[x],
)
menu = st.sidebar.radio("Ir para", ["Dre", "Notas Fiscais"])


# ==========================================
# TELA 1: DRE X CONTÁBIL
# ==========================================
if menu == "Dre":
    st.title(f"📊 DRE x Contábil - {EMPRESAS[empresa_selecionada]}")
    st.markdown("Comparativo entre o relatório contábil e a DRE do cliente, com parametrização separada por empresa.")

    aba_comparativo_dre, aba_parametrizacao_dre = st.tabs(["Comparativo", "Parametrização"])

    with aba_comparativo_dre:
        col1, col2 = st.columns(2)
        with col1:
            file_contabil = st.file_uploader("📂 Upload CONTÁBIL (.xlsx)", type=["xlsx", "xls"], key="contabil")
        with col2:
            file_cliente = st.file_uploader("📂 Upload CLIENTE / DRE (.xlsx)", type=["xlsx", "xls"], key="cliente")

        if file_contabil and file_cliente:
            if st.button("🚀 Iniciar Conferência", use_container_width=True, type="primary"):
                with st.spinner("Lendo arquivos e cruzando com o banco de dados..."):
                    try:
                        df_contabil = limpar_colunas(pd.read_excel(file_contabil, engine="openpyxl"))
                        df_cliente = limpar_colunas(pd.read_excel(file_cliente, engine="openpyxl"))

                        col_conta_contabil = obter_coluna(df_contabil, ["codic"])
                        col_grupo_contabil = obter_coluna(df_contabil, ["nomec"])
                        col_valor_contabil = obter_coluna(df_contabil, ["valdeb"])
                        col_historico_contabil = obter_coluna(df_contabil, ["historico_excel", "historico"])
                        col_tipo_lan_contabil = obter_coluna(df_contabil, ["tipo_lan"])

                        col_plano_01_cliente = obter_coluna(df_cliente, ["Plano de conta nº. 01"])
                        col_conta_cliente = obter_coluna(df_cliente, ["Cód. Plano de conta nº. 04"])
                        col_nivel3_cliente = obter_coluna(df_cliente, ["Plano de conta nº. 03"])
                        col_valor_cliente = obter_coluna(df_cliente, ["Débito"])
                        col_descricao_cliente = obter_coluna(df_cliente, ["Descrição"])
                        col_fornecedor_cliente = obter_coluna(df_cliente, ["Cliente/Fornecedor"])

                        faltantes = []
                        if not col_conta_contabil:
                            faltantes.append("codic")
                        if not col_grupo_contabil:
                            faltantes.append("nomec")
                        if not col_valor_contabil:
                            faltantes.append("valdeb")
                        if not col_conta_cliente:
                            faltantes.append("Cód. Plano de conta nº. 04")
                        if not col_nivel3_cliente:
                            faltantes.append("Plano de conta nº. 03")
                        if not col_valor_cliente:
                            faltantes.append("Débito")

                        if faltantes:
                            raise ValueError(
                                "As planilhas enviadas não têm as colunas esperadas: " + ", ".join(faltantes)
                            )

                        df_contabil["Valor_Tratado"] = df_contabil[col_valor_contabil].apply(tratar_valor)
                        df_contabil["_Conta_Contabil"] = df_contabil[col_conta_contabil].apply(normalizar_chave)
                        df_contabil["_Grupo_Contabil"] = df_contabil[col_grupo_contabil].apply(normalizar_texto)

                        df_cliente["Valor_Tratado"] = df_cliente[col_valor_cliente].apply(tratar_valor)
                        df_cliente["_Nivel_3"] = df_cliente[col_nivel3_cliente].apply(normalizar_texto)
                        df_cliente["_Nivel_4"] = df_cliente[col_conta_cliente].apply(normalizar_texto)
                        df_cliente["_Nivel_4_Norm"] = df_cliente["_Nivel_4"].apply(normalizar_chave_relacionamento)
                        if col_fornecedor_cliente:
                            df_cliente["_Fornecedor"] = df_cliente[col_fornecedor_cliente].apply(normalizar_texto)

                        antes_contabil = len(df_contabil)
                        antes_cliente = len(df_cliente)
                        df_contabil = aplicar_filtros_contabil(df_contabil, col_grupo_contabil, col_tipo_lan_contabil)
                        df_cliente = aplicar_filtros_cliente(df_cliente, col_plano_01_cliente, col_conta_cliente)
                        removidos_contabil = antes_contabil - len(df_contabil)
                        removidos_cliente = antes_cliente - len(df_cliente)
                        if removidos_contabil or removidos_cliente:
                            st.info(
                                f"Linhas desconsideradas nesta análise: {removidos_contabil} no contábil e {removidos_cliente} no cliente."
                            )

                        response = (
                            supabase.table("parametrizacao_contas")
                            .select("*")
                            .eq("empresa_id", str(empresa_selecionada))
                            .execute()
                        )
                        df_parametros = pd.DataFrame(response.data)

                        if df_parametros.empty:
                            st.warning(
                                f"⚠️ Nenhuma regra encontrada no banco para a empresa {EMPRESAS[empresa_selecionada]}. Vá na aba Parametrização!"
                            )
                        else:
                            df_parametros = limpar_colunas(df_parametros)
                            if "conta_contabil" not in df_parametros.columns or "fornecedor_cliente" not in df_parametros.columns:
                                raise ValueError("A tabela parametrizacao_contas precisa ter as colunas conta_contabil e fornecedor_cliente.")

                            df_parametros["conta_contabil"] = df_parametros["conta_contabil"].apply(normalizar_chave)
                            df_parametros["fornecedor_cliente"] = df_parametros["fornecedor_cliente"].apply(
                                normalizar_chave_relacionamento
                            )

                            df_cont_grp = (
                                df_contabil.groupby("_Conta_Contabil")
                                .agg(
                                    Grupo_Contabil=("_Grupo_Contabil", "first"),
                                    Valor_Contabil=("Valor_Tratado", "sum"),
                                    Qtde_Contabil=("_Conta_Contabil", "size"),
                                )
                                .reset_index()
                                .rename(columns={"_Conta_Contabil": "Conta Débito"})
                            )

                            resultados = []
                            detalhes_por_conta = {}

                            for _, row in df_cont_grp.iterrows():
                                conta_contabil = str(row["Conta Débito"]).strip()
                                valor_contabil = round(float(row["Valor_Contabil"]), 2)
                                grupo_contabil = str(row["Grupo_Contabil"]).strip()
                                qtde_contabil = int(row["Qtde_Contabil"])

                                if valor_contabil == 0:
                                    continue

                                regras = df_parametros[df_parametros["conta_contabil"] == conta_contabil]
                                df_cont_detalhe = df_contabil[df_contabil["_Conta_Contabil"] == conta_contabil].copy()

                                if regras.empty:
                                    resultados.append(
                                        {
                                            "Conta Débito": conta_contabil,
                                            "Grupo DRE (Contábil)": grupo_contabil,
                                            "Valor Contábil": valor_contabil,
                                            "Valor Cliente": 0.0,
                                            "Diferença": valor_contabil,
                                            "QTD CONTÁBIL": qtde_contabil,
                                            "QTD CLIENTE": 0,
                                            "DIF. QTD": qtde_contabil,
                                            "STATUS": "⚠️ Sem parametrização",
                                            "MOTIVO": "Conta sem regra cadastrada no mini banco.",
                                        }
                                    )
                                    detalhes_por_conta[conta_contabil] = {
                                        "contabil": df_cont_detalhe,
                                        "cliente": pd.DataFrame(),
                                    }
                                else:
                                    lista_fornecedores = set(regras["fornecedor_cliente"].tolist())
                                    filtro_cli = df_cliente["_Nivel_4_Norm"].isin(lista_fornecedores)
                                    df_cli_match = df_cliente[filtro_cli].copy()

                                    valor_cliente = round(df_cli_match["Valor_Tratado"].sum(), 2)
                                    qtde_cliente = int(len(df_cli_match))
                                    diff = round(valor_contabil - valor_cliente, 2)
                                    diff_qtd = qtde_contabil - qtde_cliente
                                    status, motivo = classificar_diferenca(diff)

                                    resultados.append(
                                        {
                                            "Conta Débito": conta_contabil,
                                            "Grupo DRE (Contábil)": grupo_contabil,
                                            "Valor Contábil": valor_contabil,
                                            "Valor Cliente": valor_cliente,
                                            "Diferença": diff,
                                            "QTD CONTÁBIL": qtde_contabil,
                                            "QTD CLIENTE": qtde_cliente,
                                            "DIF. QTD": diff_qtd,
                                            "STATUS": status,
                                            "MOTIVO": motivo,
                                        }
                                    )

                                    detalhes_por_conta[conta_contabil] = {
                                        "contabil": df_cont_detalhe,
                                        "cliente": df_cli_match,
                                    }

                            df_resultados = pd.DataFrame(resultados)
                            assinatura_atual = assinatura_arquivos(file_contabil, file_cliente)
                            st.session_state["analise_conciliacao"] = {
                                "df_resultados": df_resultados,
                                "detalhes_por_conta": detalhes_por_conta,
                                "empresa_id": empresa_selecionada,
                                "empresa_nome": EMPRESAS[empresa_selecionada],
                                "df_cliente": df_cliente,
                                "df_contabil": df_contabil,
                                "col_conta_contabil": col_conta_contabil,
                                "col_grupo_contabil": col_grupo_contabil,
                                "col_historico_contabil": col_historico_contabil,
                                "col_nivel3_cliente": col_nivel3_cliente,
                                "col_conta_cliente": col_conta_cliente,
                                "col_fornecedor_cliente": col_fornecedor_cliente,
                                "col_descricao_cliente": col_descricao_cliente,
                            }
                            st.session_state["analise_conciliacao_assinatura"] = assinatura_atual
                            st.session_state["analise_conciliacao_empresa"] = empresa_selecionada

                    except Exception as e:
                        st.error(f"❌ Erro ao ler planilhas ou cruzar dados. Detalhe: {e}")

        analise_armazenada = st.session_state.get("analise_conciliacao")
        assinatura_atual = assinatura_arquivos(file_contabil, file_cliente) if file_contabil and file_cliente else None
        if (
            analise_armazenada
            and st.session_state.get("analise_conciliacao_empresa") == empresa_selecionada
            and st.session_state.get("analise_conciliacao_assinatura") == assinatura_atual
        ):
            renderizar_analise_conferencia(analise_armazenada)
        elif file_contabil and file_cliente:
            st.info("Clique em Iniciar Conferência para gerar a análise.")

    with aba_parametrizacao_dre:
        st.subheader("Parametrização")
        st.caption("Aqui você consulta o que já foi parametrizado e identifica o que ainda falta vincular no cliente.")

        df_banco = carregar_parametrizacao_empresa(empresa_selecionada)

        st.subheader("Consulta de Parametrização")
        st.caption("Conta Débito do contábil x Cód. Plano de conta nº. 04 do cliente.")

        filtro_consulta = st.text_input("Buscar na parametrização", placeholder="Digite conta débito ou código do cliente...")
        if filtro_consulta:
            termo = filtro_consulta.strip().lower()
            mascara = (
                df_banco["conta_contabil"].astype(str).str.lower().str.contains(termo, na=False)
                | df_banco["fornecedor_cliente"].astype(str).str.lower().str.contains(termo, na=False)
            )
            df_banco_visivel = df_banco[mascara].copy()
        else:
            df_banco_visivel = df_banco.copy()

        if df_banco_visivel.empty:
            st.info("Nenhuma regra encontrada para o filtro atual.")
        else:
            st.dataframe(
                df_banco_visivel.rename(
                    columns={
                        "conta_contabil": "CONTA DÉBITO",
                        "fornecedor_cliente": "CÓD. PLANO 04",
                    }
                )[["CONTA DÉBITO", "CÓD. PLANO 04"]]
                .sort_values(["CONTA DÉBITO", "CÓD. PLANO 04"]),
                use_container_width=True,
                hide_index=True,
            )

        st.markdown("---")
        st.subheader("Editar ou excluir parametrizações")
        st.caption("Altere a conta débito, marque a linha para excluir ou adicione novas linhas diretamente na tabela.")

        df_gerenciar = df_banco.copy()
        if df_gerenciar.empty:
            df_gerenciar = pd.DataFrame(columns=["id", "conta_contabil", "fornecedor_cliente", "EXCLUIR"])
        else:
            df_gerenciar["EXCLUIR"] = False

        with st.form("form_gerenciar_parametrizacao"):
            tabela_gerenciar = st.data_editor(
                df_gerenciar,
                num_rows="dynamic",
                use_container_width=True,
                hide_index=True,
                column_config={
                    "id": st.column_config.TextColumn("ID Supabase", disabled=True),
                    "conta_contabil": st.column_config.TextColumn("Conta Débito", required=True),
                    "fornecedor_cliente": st.column_config.TextColumn("Cód. Plano de conta nº. 04", required=True),
                    "EXCLUIR": st.column_config.CheckboxColumn("Excluir", help="Marque para remover este vínculo."),
                },
            )

            salvar_gerenciamento = st.form_submit_button("💾 Salvar alterações da parametrização", type="primary")

            if salvar_gerenciamento:
                try:
                    dados_salvar = []
                    ids_excluir = []
                    vistos = set()

                    for _, row in tabela_gerenciar.iterrows():
                        row_id = row.get("id")
                        excluir = bool(row.get("EXCLUIR", False))
                        conta = normalizar_chave(row.get("conta_contabil", ""))
                        fornecedor = normalizar_texto(row.get("fornecedor_cliente", ""))

                        if pd.notna(row_id):
                            row_id = str(row_id)

                        if excluir:
                            if row_id:
                                ids_excluir.append(row_id)
                            elif fornecedor:
                                ids_excluir.append(f"{empresa_selecionada}:{fornecedor}")
                            continue

                        if not conta or not fornecedor:
                            continue

                        chave = (str(empresa_selecionada), fornecedor)
                        if chave in vistos:
                            continue
                        vistos.add(chave)

                        dados_salvar.append(
                            {
                                "empresa_id": str(empresa_selecionada),
                                "conta_contabil": conta,
                                "fornecedor_cliente": fornecedor,
                            }
                        )

                    for item in ids_excluir:
                        if ":" in item:
                            _, fornecedor = item.split(":", 1)
                            supabase.table("parametrizacao_contas").delete() \
                                .eq("empresa_id", str(empresa_selecionada)) \
                                .eq("fornecedor_cliente", fornecedor) \
                                .execute()
                        else:
                            supabase.table("parametrizacao_contas").delete().eq("id", item).execute()

                    if dados_salvar:
                        try:
                            supabase.table("parametrizacao_contas").upsert(
                                dados_salvar,
                                on_conflict="empresa_id,fornecedor_cliente",
                            ).execute()
                        except Exception:
                            for item in dados_salvar:
                                supabase.table("parametrizacao_contas").delete() \
                                    .eq("empresa_id", item["empresa_id"]) \
                                    .eq("fornecedor_cliente", item["fornecedor_cliente"]) \
                                    .execute()
                            supabase.table("parametrizacao_contas").insert(dados_salvar).execute()

                    st.success("✅ Parametrizações atualizadas com sucesso!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao atualizar parametrizações: {e}")

        st.markdown("---")
        st.subheader("Fornecedores sem parametrização")
        st.caption("Lista automática dos lançamentos do cliente que ainda não têm de/para cadastrado.")

        fonte_cliente = st.session_state.get("analise_conciliacao", {}).get("df_cliente")
        col_nivel3_cliente = st.session_state.get("analise_conciliacao", {}).get("col_nivel3_cliente")
        col_conta_cliente = st.session_state.get("analise_conciliacao", {}).get("col_conta_cliente")
        col_fornecedor_cliente = st.session_state.get("analise_conciliacao", {}).get("col_fornecedor_cliente")
        col_descricao_cliente = st.session_state.get("analise_conciliacao", {}).get("col_descricao_cliente")

        if fonte_cliente is None or fonte_cliente.empty:
            st.info(
                "Para listar os fornecedores sem parametrização, carregue uma base do cliente abaixo ou rode uma conciliação primeiro."
            )
            arquivo_cliente_consulta = st.file_uploader(
                "📂 Upload CLIENTE / DRE para consulta",
                type=["xlsx", "xls"],
                key="cliente_consulta_parametrizacao",
            )
            if arquivo_cliente_consulta:
                fonte_cliente = limpar_colunas(pd.read_excel(arquivo_cliente_consulta, engine="openpyxl"))
                col_plano_01_cliente = obter_coluna(fonte_cliente, ["Plano de conta nº. 01"])
                col_nivel3_cliente = obter_coluna(fonte_cliente, ["Plano de conta nº. 03"])
                col_conta_cliente = obter_coluna(fonte_cliente, ["Cód. Plano de conta nº. 04"])
                col_fornecedor_cliente = obter_coluna(fonte_cliente, ["Cliente/Fornecedor"])
                col_descricao_cliente = obter_coluna(fonte_cliente, ["Descrição"])
                col_valor_cliente = obter_coluna(fonte_cliente, ["Débito"])
                if col_nivel3_cliente and col_conta_cliente and col_valor_cliente:
                    fonte_cliente = aplicar_filtros_cliente(fonte_cliente, col_plano_01_cliente, col_conta_cliente)
                    fonte_cliente["Valor_Tratado"] = fonte_cliente[col_valor_cliente].apply(tratar_valor)
                    fonte_cliente["_Nivel_3"] = fonte_cliente[col_nivel3_cliente].apply(normalizar_texto)
                    fonte_cliente["_Nivel_4"] = fonte_cliente[col_conta_cliente].apply(normalizar_texto)
                    if col_fornecedor_cliente:
                        fonte_cliente["_Fornecedor"] = fonte_cliente[col_fornecedor_cliente].apply(normalizar_texto)
                else:
                    st.warning("A base enviada não tem as colunas mínimas esperadas para consulta.")

        if fonte_cliente is not None and not fonte_cliente.empty and col_nivel3_cliente and col_conta_cliente:
            df_pendencias = preparar_pendencias_parametrizacao(
                fonte_cliente,
                df_banco,
                col_nivel3_cliente,
                col_conta_cliente,
                col_fornecedor_cliente,
                col_descricao_cliente,
            )

            if df_pendencias.empty:
                st.success("Todas as contas/fornecedores dessa base já estão parametrizados.")
            else:
                st.subheader("Montar parametrização das pendências")
                st.caption("Preencha a conta débito nas linhas pendentes e salve para criar os vínculos no banco.")

                df_edicao = df_pendencias.copy()
                df_edicao.insert(0, "CONTA DÉBITO", "")

                with st.form("form_param_pendencias"):
                    tabela_parametrizacao = st.data_editor(
                        df_edicao,
                        num_rows="dynamic",
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "CONTA DÉBITO": st.column_config.TextColumn("CONTA DÉBITO", required=False),
                            "GRUPO DRE": st.column_config.TextColumn("GRUPO DRE", disabled=True),
                            "CÓD. PLANO 04": st.column_config.TextColumn("CÓD. PLANO 04", disabled=True),
                            "FORNECEDOR": st.column_config.TextColumn("FORNECEDOR", disabled=True),
                            "VALOR": st.column_config.NumberColumn("VALOR", disabled=True, format="R$ %.2f"),
                            "QTD LANCAMENTOS": st.column_config.NumberColumn("QTD LANCAMENTOS", disabled=True),
                            "STATUS": st.column_config.TextColumn("STATUS", disabled=True),
                        },
                    )

                    salvar_parametrizacao = st.form_submit_button("💾 Salvar parametrizações pendentes", type="primary")

                    if salvar_parametrizacao:
                        try:
                            novos_dados = []
                            vistos = set()

                            for _, row in tabela_parametrizacao.iterrows():
                                conta = normalizar_chave(row.get("CONTA DÉBITO", ""))
                                fornecedor = normalizar_texto(row.get("CÓD. PLANO 04", ""))

                                if not conta or not fornecedor:
                                    continue

                                chave = (str(empresa_selecionada), conta, fornecedor)
                                if chave in vistos:
                                    continue
                                vistos.add(chave)

                                novos_dados.append(
                                    {
                                        "empresa_id": str(empresa_selecionada),
                                        "conta_contabil": conta,
                                        "fornecedor_cliente": fornecedor,
                                    }
                                )

                            if novos_dados:
                                try:
                                    supabase.table("parametrizacao_contas").upsert(
                                        novos_dados,
                                        on_conflict="empresa_id,fornecedor_cliente",
                                    ).execute()
                                except Exception:
                                    for item in novos_dados:
                                        supabase.table("parametrizacao_contas").delete() \
                                            .eq("empresa_id", item["empresa_id"]) \
                                            .eq("fornecedor_cliente", item["fornecedor_cliente"]) \
                                            .execute()
                                    supabase.table("parametrizacao_contas").insert(novos_dados).execute()

                            st.success("✅ Parametrizações salvas com sucesso!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erro ao salvar parametrizações: {e}")
        elif fonte_cliente is not None and fonte_cliente.empty:
            st.warning("Não foi possível montar a lista de pendências porque a base do cliente está vazia.")


# ==========================================
# TELA 3: NOTAS FISCAIS X CONTÁBIL
# ==========================================
elif menu == "Notas Fiscais":
    st.title(f"🧾 Notas Fiscais x Razão - {EMPRESAS[empresa_selecionada]}")
    st.markdown("Comparativo entre o razão contábil e os relatórios de NFe/NFSe, com banco de parametrização separado.")

    aba_comparativo, aba_parametrizacao = st.tabs(["Comparativo", "Parametrização"])

    with aba_comparativo:
        c1, c2, c3 = st.columns(3)
        with c1:
            file_contabil_docs = st.file_uploader("📂 Upload RAZÃO CONTÁBIL (.xlsx)", type=["xlsx", "xls"], key="contabil_documentos")
        with c2:
            file_nfse = st.file_uploader("📂 Upload NFSe (.xlsx)", type=["xlsx", "xls"], key="nfse_documentos")
        with c3:
            file_nfe = st.file_uploader("📂 Upload NFe (.xlsx)", type=["xlsx", "xls"], key="nfe_documentos")

        if file_contabil_docs or file_nfse or file_nfe:
            try:
                totais_relatorios, avisos_totais = calcular_totais_relatorios_notas(
                    file_contabil_docs,
                    file_nfse,
                    file_nfe,
                )

                st.markdown("### Quadro Geral de Totais")
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Contábil", formatar_moeda(totais_relatorios["contabil"]))
                m2.metric("NFSe", formatar_moeda(totais_relatorios["nfse"]))
                m3.metric("NFe", formatar_moeda(totais_relatorios["nfe"]))
                m4.metric("NFSe + NFe", formatar_moeda(totais_relatorios["documentos"]))
                st.caption("Resumo rápido dos valores totais dos arquivos enviados. No contábil, o total considera apenas `tipo_lan = D` quando essa coluna existir.")

                for aviso_total in avisos_totais:
                    st.warning(aviso_total)
            except Exception as e:
                st.warning(f"Não foi possível montar o quadro geral de totais: {e}")

        if file_contabil_docs and (file_nfse or file_nfe):
            if st.button("🚀 Iniciar Comparação de Notas", use_container_width=True, type="primary"):
                with st.spinner("Lendo arquivos e cruzando com o banco de parametrização das notas..."):
                    try:
                        df_contabil = limpar_colunas(pd.read_excel(file_contabil_docs, engine="openpyxl"))
                        df_nfse = limpar_colunas(pd.read_excel(file_nfse, engine="openpyxl")) if file_nfse else pd.DataFrame()
                        df_nfe = limpar_colunas(pd.read_excel(file_nfe, engine="openpyxl")) if file_nfe else pd.DataFrame()

                        col_conta_contabil = obter_coluna_flexivel(df_contabil, ["codic"])
                        col_grupo_contabil = obter_coluna_flexivel(df_contabil, ["nomec"])
                        col_valor_contabil = obter_coluna_flexivel(df_contabil, ["valdeb"])
                        col_historico_contabil = obter_coluna_flexivel(df_contabil, ["historico_excel", "historico"])
                        col_tipo_lan_contabil = obter_coluna_flexivel(df_contabil, ["tipo_lan"])

                        faltantes = []
                        if not col_conta_contabil:
                            faltantes.append("codic")
                        if not col_grupo_contabil:
                            faltantes.append("nomec")
                        if not col_valor_contabil:
                            faltantes.append("valdeb")
                        if faltantes:
                            raise ValueError("O razão contábil não tem as colunas esperadas: " + ", ".join(faltantes))

                        df_contabil["Valor_Tratado"] = df_contabil[col_valor_contabil].apply(tratar_valor)
                        df_contabil["_Conta_Contabil"] = df_contabil[col_conta_contabil].apply(normalizar_chave)
                        df_contabil["_Grupo_Contabil"] = df_contabil[col_grupo_contabil].apply(normalizar_texto)
                        if col_tipo_lan_contabil:
                            df_contabil = df_contabil[
                                df_contabil[col_tipo_lan_contabil].apply(normalizar_texto).str.upper() == "D"
                            ].copy()

                        df_documentos, avisos_documentos = preparar_base_documentos_fiscais(df_nfe, df_nfse)
                        for aviso in avisos_documentos:
                            st.warning(aviso)

                        if df_documentos.empty:
                            raise ValueError("Nenhum lançamento válido de NFe/NFSe foi encontrado para comparar.")

                        df_parametros = carregar_parametrizacao_notas_empresa(empresa_selecionada)
                        erro_tabela_notas = st.session_state.get("erro_parametrizacao_notas")
                        if erro_tabela_ausente(erro_tabela_notas):
                            st.error(
                                f"A tabela `{TABELA_PARAMETRIZACAO_NOTAS}` ainda não existe no Supabase. Ela é separada da parametrização antiga."
                            )
                            st.code(sql_tabela_parametrizacao_notas(), language="sql")
                        if df_parametros.empty:
                            st.warning(
                                f"⚠️ Nenhuma regra encontrada no banco de notas para a empresa {EMPRESAS[empresa_selecionada]}. Vá na aba Parametrização."
                            )
                        else:
                            df_parametros = limpar_colunas(df_parametros)
                            if "conta_contabil" not in df_parametros.columns or "codigo_plano_04" not in df_parametros.columns:
                                raise ValueError(
                                    f"A tabela {TABELA_PARAMETRIZACAO_NOTAS} precisa ter as colunas conta_contabil e codigo_plano_04."
                                )

                            df_parametros["conta_contabil"] = df_parametros["conta_contabil"].apply(normalizar_chave)
                            df_parametros["codigo_plano_04"] = df_parametros["codigo_plano_04"].apply(normalizar_chave_relacionamento)

                            df_cont_grp = (
                                df_contabil.groupby("_Conta_Contabil")
                                .agg(
                                    Grupo_Contabil=("_Grupo_Contabil", "first"),
                                    Valor_Contabil=("Valor_Tratado", "sum"),
                                    Qtde_Contabil=("_Conta_Contabil", "size"),
                                )
                                .reset_index()
                                .rename(columns={"_Conta_Contabil": "Conta Débito"})
                            )

                            resultados = []
                            detalhes_por_conta = {}

                            for _, row in df_cont_grp.iterrows():
                                conta_contabil = str(row["Conta Débito"]).strip()
                                valor_contabil = round(float(row["Valor_Contabil"]), 2)
                                grupo_contabil = str(row["Grupo_Contabil"]).strip()
                                qtde_contabil = int(row["Qtde_Contabil"])

                                if valor_contabil == 0:
                                    continue

                                regras = df_parametros[df_parametros["conta_contabil"] == conta_contabil]
                                df_cont_detalhe = df_contabil[df_contabil["_Conta_Contabil"] == conta_contabil].copy()

                                if regras.empty:
                                    resultados.append(
                                        {
                                            "Conta Débito": conta_contabil,
                                            "Grupo DRE (Contábil)": grupo_contabil,
                                            "Valor Contábil": valor_contabil,
                                            "Valor Documentos": 0.0,
                                            "Diferença": valor_contabil,
                                            "QTD CONTÁBIL": qtde_contabil,
                                            "QTD DOCUMENTOS": 0,
                                            "STATUS": "⚠️ Sem parametrização",
                                            "MOTIVO": "Conta sem regra cadastrada no banco de notas.",
                                        }
                                    )
                                    detalhes_por_conta[conta_contabil] = {"contabil": df_cont_detalhe, "documentos": pd.DataFrame()}
                                else:
                                    codigos_04 = set(regras["codigo_plano_04"].tolist())
                                    df_docs_match = df_documentos[df_documentos["_Nivel_4_Norm"].isin(codigos_04)].copy()
                                    valor_documentos = round(df_docs_match["Valor_Tratado"].sum(), 2)
                                    qtde_documentos = int(len(df_docs_match))
                                    diff = round(valor_contabil - valor_documentos, 2)
                                    status, motivo = classificar_diferenca(diff)

                                    resultados.append(
                                        {
                                            "Conta Débito": conta_contabil,
                                            "Grupo DRE (Contábil)": grupo_contabil,
                                            "Valor Contábil": valor_contabil,
                                            "Valor Documentos": valor_documentos,
                                            "Diferença": diff,
                                            "QTD CONTÁBIL": qtde_contabil,
                                            "QTD DOCUMENTOS": qtde_documentos,
                                            "STATUS": status,
                                            "MOTIVO": motivo,
                                        }
                                    )
                                    detalhes_por_conta[conta_contabil] = {"contabil": df_cont_detalhe, "documentos": df_docs_match}

                            st.session_state["analise_documentos_fiscais"] = {
                                "df_resultados": pd.DataFrame(resultados),
                                "detalhes_por_conta": detalhes_por_conta,
                                "empresa_id": empresa_selecionada,
                                "empresa_nome": EMPRESAS[empresa_selecionada],
                                "df_documentos": df_documentos,
                                "col_conta_contabil": col_conta_contabil,
                                "col_grupo_contabil": col_grupo_contabil,
                                "col_historico_contabil": col_historico_contabil,
                            }
                            st.session_state["analise_documentos_assinatura"] = assinatura_arquivos(
                                file_contabil_docs, file_nfse, file_nfe
                            )
                            st.session_state["analise_documentos_empresa"] = empresa_selecionada

                    except Exception as e:
                        st.error(f"❌ Erro ao comparar razão com NFe/NFSe. Detalhe: {e}")

        analise_documentos = st.session_state.get("analise_documentos_fiscais")
        assinatura_docs = assinatura_arquivos(file_contabil_docs, file_nfse, file_nfe) if file_contabil_docs and (file_nfse or file_nfe) else None
        if (
            analise_documentos
            and st.session_state.get("analise_documentos_empresa") == empresa_selecionada
            and st.session_state.get("analise_documentos_assinatura") == assinatura_docs
        ):
            renderizar_analise_documentos(analise_documentos)
        elif file_contabil_docs and (file_nfse or file_nfe):
            st.info("Clique em Iniciar Comparação de Notas para gerar a análise.")

    with aba_parametrizacao:
        st.subheader("Parametrização das Notas")
        st.caption("Banco separado da conciliação antiga: Nível 03 + Nível 04 do cliente vinculados à conta contábil.")

        df_banco_notas = carregar_parametrizacao_notas_empresa(empresa_selecionada)
        erro_tabela_notas = st.session_state.get("erro_parametrizacao_notas")
        if erro_tabela_ausente(erro_tabela_notas):
            st.error(
                f"A tabela `{TABELA_PARAMETRIZACAO_NOTAS}` ainda não existe no Supabase. Eu não mexi no banco; essa estrutura precisa ser criada manualmente."
            )
            st.code(sql_tabela_parametrizacao_notas(), language="sql")
        df_seed_notas = carregar_seed_parametrizacao_notas()
        tem_seed_empresa = not df_seed_notas[df_seed_notas["empresa_id"].astype(str) == str(empresa_selecionada)].empty

        if tem_seed_empresa:
            if st.button("📥 Carregar base padrão desta empresa no banco de notas", use_container_width=True):
                try:
                    dados_seed = df_seed_notas[df_seed_notas["empresa_id"].astype(str) == str(empresa_selecionada)].copy()
                    registros = []
                    for _, row in dados_seed.iterrows():
                        codigo_04 = normalizar_texto(row.get("codigo_plano_04", ""))
                        if not codigo_04:
                            continue
                        registros.append(
                            {
                                "empresa_id": str(empresa_selecionada),
                                "plano_conta_nivel_03": normalizar_texto(row.get("plano_conta_nivel_03", "")),
                                "codigo_plano_04": codigo_04,
                                "conta_contabil": normalizar_chave(row.get("conta_contabil", "")),
                            }
                        )
                    if registros:
                        supabase.table(TABELA_PARAMETRIZACAO_NOTAS).upsert(
                            registros,
                            on_conflict="empresa_id,codigo_plano_04",
                        ).execute()
                    st.success("✅ Base padrão carregada no banco de notas.")
                    st.rerun()
                except Exception as e:
                    if erro_tabela_ausente(e):
                        st.error(
                            f"A tabela `{TABELA_PARAMETRIZACAO_NOTAS}` ainda não existe no Supabase. Rode o SQL acima e tente de novo."
                        )
                    elif erro_rls_policy(e):
                        st.error(
                            "O Supabase bloqueou a gravação por RLS. Para salvar pela aplicação, configure `SUPABASE_SERVICE_ROLE_KEY` no Streamlit Cloud ou crie uma policy de insert/update para essa tabela."
                        )
                    else:
                        st.error(f"Erro ao carregar base padrão: {e}")

        filtro_consulta_notas = st.text_input(
            "Buscar na parametrização de notas",
            placeholder="Digite nível 03, nível 04 ou conta contábil...",
            key="filtro_param_notas",
        )
        if filtro_consulta_notas:
            termo = filtro_consulta_notas.strip().lower()
            mascara = (
                df_banco_notas["plano_conta_nivel_03"].astype(str).str.lower().str.contains(termo, na=False)
                | df_banco_notas["codigo_plano_04"].astype(str).str.lower().str.contains(termo, na=False)
                | df_banco_notas["conta_contabil"].astype(str).str.lower().str.contains(termo, na=False)
            )
            df_banco_notas_visivel = df_banco_notas[mascara].copy()
        else:
            df_banco_notas_visivel = df_banco_notas.copy()

        if df_banco_notas_visivel.empty:
            st.info("Nenhuma regra encontrada para o filtro atual.")
        else:
            st.dataframe(
                df_banco_notas_visivel.rename(
                    columns={
                        "plano_conta_nivel_03": "PLANO DE CONTA Nº. 03",
                        "codigo_plano_04": "CÓD. PLANO DE CONTA Nº. 04",
                        "conta_contabil": "CONTA CONTÁBIL",
                    }
                )[["PLANO DE CONTA Nº. 03", "CÓD. PLANO DE CONTA Nº. 04", "CONTA CONTÁBIL"]],
                use_container_width=True,
                hide_index=True,
            )

        st.markdown("---")
        st.subheader("Editar ou excluir parametrizações das notas")

        df_gerenciar_notas = df_banco_notas.copy()
        if df_gerenciar_notas.empty:
            df_gerenciar_notas = pd.DataFrame(
                columns=["id", "plano_conta_nivel_03", "codigo_plano_04", "conta_contabil", "EXCLUIR"]
            )
        else:
            df_gerenciar_notas["EXCLUIR"] = False

        with st.form("form_gerenciar_parametrizacao_notas"):
            tabela_gerenciar_notas = st.data_editor(
                df_gerenciar_notas,
                num_rows="dynamic",
                use_container_width=True,
                hide_index=True,
                column_config={
                    "id": st.column_config.TextColumn("ID Supabase", disabled=True),
                    "plano_conta_nivel_03": st.column_config.TextColumn("Plano de conta nº. 03", required=False),
                    "codigo_plano_04": st.column_config.TextColumn("Cód. Plano de conta nº. 04", required=True),
                    "conta_contabil": st.column_config.TextColumn("Conta contábil", required=True),
                    "EXCLUIR": st.column_config.CheckboxColumn("Excluir"),
                },
            )
            salvar_gerenciamento_notas = st.form_submit_button("💾 Salvar alterações das notas", type="primary")

            if salvar_gerenciamento_notas:
                try:
                    dados_salvar = []
                    ids_excluir = []
                    vistos = set()

                    for _, row in tabela_gerenciar_notas.iterrows():
                        row_id = row.get("id")
                        excluir = bool(row.get("EXCLUIR", False))
                        plano_03 = normalizar_texto(row.get("plano_conta_nivel_03", ""))
                        codigo_04 = normalizar_texto(row.get("codigo_plano_04", ""))
                        conta = normalizar_chave(row.get("conta_contabil", ""))

                        if pd.notna(row_id):
                            row_id = str(row_id)

                        if excluir:
                            if row_id:
                                ids_excluir.append(row_id)
                            elif codigo_04:
                                ids_excluir.append(f"{empresa_selecionada}:{codigo_04}")
                            continue

                        if not codigo_04 or not conta:
                            continue

                        chave = (str(empresa_selecionada), codigo_04)
                        if chave in vistos:
                            continue
                        vistos.add(chave)

                        dados_salvar.append(
                            {
                                "empresa_id": str(empresa_selecionada),
                                "plano_conta_nivel_03": plano_03,
                                "codigo_plano_04": codigo_04,
                                "conta_contabil": conta,
                            }
                        )

                    for item in ids_excluir:
                        if ":" in item:
                            _, codigo_04 = item.split(":", 1)
                            supabase.table(TABELA_PARAMETRIZACAO_NOTAS).delete() \
                                .eq("empresa_id", str(empresa_selecionada)) \
                                .eq("codigo_plano_04", codigo_04) \
                                .execute()
                        else:
                            supabase.table(TABELA_PARAMETRIZACAO_NOTAS).delete().eq("id", item).execute()

                    if dados_salvar:
                        supabase.table(TABELA_PARAMETRIZACAO_NOTAS).upsert(
                            dados_salvar,
                            on_conflict="empresa_id,codigo_plano_04",
                        ).execute()

                    st.success("✅ Parametrizações das notas atualizadas com sucesso!")
                    st.rerun()
                except Exception as e:
                    if erro_tabela_ausente(e):
                        st.error(
                            f"A tabela `{TABELA_PARAMETRIZACAO_NOTAS}` ainda não existe no Supabase. Rode o SQL acima e tente de novo."
                        )
                    elif erro_rls_policy(e):
                        st.error(
                            "O Supabase bloqueou a gravação por RLS. Para salvar pela aplicação, configure `SUPABASE_SERVICE_ROLE_KEY` no Streamlit Cloud ou crie uma policy de insert/update para essa tabela."
                        )
                    else:
                        st.error(f"Erro ao atualizar parametrizações das notas: {e}")

        st.markdown("---")
        st.subheader("Planos sem parametrização")
        st.caption("Lista única dos planos nível 04 encontrados em NFe/NFSe e ainda não cadastrados no banco novo.")

        fonte_documentos = st.session_state.get("analise_documentos_fiscais", {}).get("df_documentos")
        if fonte_documentos is None or fonte_documentos.empty:
            c1, c2 = st.columns(2)
            with c1:
                arquivo_nfse_param = st.file_uploader("📂 Upload NFSe para parametrização", type=["xlsx", "xls"], key="nfse_param")
            with c2:
                arquivo_nfe_param = st.file_uploader("📂 Upload NFe para parametrização", type=["xlsx", "xls"], key="nfe_param")

            if arquivo_nfse_param or arquivo_nfe_param:
                df_nfse_param = limpar_colunas(pd.read_excel(arquivo_nfse_param, engine="openpyxl")) if arquivo_nfse_param else pd.DataFrame()
                df_nfe_param = limpar_colunas(pd.read_excel(arquivo_nfe_param, engine="openpyxl")) if arquivo_nfe_param else pd.DataFrame()
                fonte_documentos, avisos_docs_param = preparar_base_documentos_fiscais(df_nfe_param, df_nfse_param)
                for aviso in avisos_docs_param:
                    st.warning(aviso)

        if fonte_documentos is not None and not fonte_documentos.empty:
            df_pendencias_docs = preparar_pendencias_documentos(fonte_documentos, df_banco_notas)
            if df_pendencias_docs.empty:
                st.success("Todos os planos desta base já estão parametrizados no banco de notas.")
            else:
                df_edicao_docs = df_pendencias_docs.copy()
                df_edicao_docs.insert(0, "CONTA CONTÁBIL", "")
                nome_empresa = EMPRESAS.get(str(empresa_selecionada), str(empresa_selecionada))
                arquivo_modelo = exportar_modelo_parametrizacao_notas(nome_empresa, df_edicao_docs)

                col_download, col_upload = st.columns([1, 1])
                with col_download:
                    st.download_button(
                        "⬇️ Baixar planilha modelo",
                        data=arquivo_modelo,
                        file_name=f"modelo_parametrizacao_notas_{empresa_selecionada}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                with col_upload:
                    with st.form(f"form_importar_modelo_notas_{empresa_selecionada}"):
                        arquivo_modelo_preenchido = st.file_uploader(
                            "📥 Importar planilha preenchida",
                            type=["xlsx", "xls"],
                            key=f"importar_modelo_notas_{empresa_selecionada}",
                        )
                        importar_modelo_notas = st.form_submit_button("📤 Importar planilha modelo")

                    if importar_modelo_notas:
                        try:
                            if not arquivo_modelo_preenchido:
                                raise ValueError("Selecione a planilha preenchida antes de importar.")
                            df_importacao = importar_modelo_parametrizacao_notas(arquivo_modelo_preenchido)
                            total_importacao = salvar_parametrizacao_notas(df_importacao, empresa_selecionada)
                            if total_importacao:
                                st.success(f"✅ {total_importacao} parametrizações das notas importadas com sucesso!")
                                st.rerun()
                            st.info("A planilha foi lida, mas não havia linhas com CONTA CONTÁBIL preenchida para importar.")
                        except Exception as e:
                            if erro_tabela_ausente(e):
                                st.error(
                                    f"A tabela `{TABELA_PARAMETRIZACAO_NOTAS}` ainda não existe no Supabase. Rode o SQL acima e tente de novo."
                                )
                            elif erro_rls_policy(e):
                                st.error(
                                    "O Supabase bloqueou a gravação por RLS. Para salvar pela aplicação, configure `SUPABASE_SERVICE_ROLE_KEY` no Streamlit Cloud ou crie uma policy de insert/update para essa tabela."
                                )
                            else:
                                st.error(f"Erro ao importar planilha de parametrização das notas: {e}")

                st.caption("Fluxo sugerido: baixe o modelo, preencha a coluna CONTA CONTÁBIL e importe a planilha para atualizar o banco.")

                with st.form("form_param_pendencias_notas"):
                    tabela_param_docs = st.data_editor(
                        df_edicao_docs,
                        num_rows="dynamic",
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "CONTA CONTÁBIL": st.column_config.TextColumn("Conta contábil", required=False),
                            "PLANO DE CONTA Nº. 03": st.column_config.TextColumn("Plano de conta nº. 03", required=False),
                            "CÓD. PLANO DE CONTA Nº. 04": st.column_config.TextColumn("Cód. Plano de conta nº. 04", disabled=True),
                            "FORNECEDOR/EMITENTE": st.column_config.TextColumn("Fornecedor/Emitente", disabled=True),
                            "ORIGENS": st.column_config.TextColumn("Origens", disabled=True),
                        },
                    )
                    salvar_pendencias_notas = st.form_submit_button("💾 Salvar parametrizações pendentes das notas", type="primary")

                    if salvar_pendencias_notas:
                        try:
                            total_salvo = salvar_parametrizacao_notas(tabela_param_docs, empresa_selecionada)
                            if total_salvo:
                                st.success(f"✅ {total_salvo} parametrizações pendentes das notas salvas com sucesso!")
                            else:
                                st.info("Nenhuma linha com CONTA CONTÁBIL preenchida foi encontrada para salvar.")
                            st.rerun()
                        except Exception as e:
                            if erro_tabela_ausente(e):
                                st.error(
                                    f"A tabela `{TABELA_PARAMETRIZACAO_NOTAS}` ainda não existe no Supabase. Rode o SQL acima e tente de novo."
                                )
                            elif erro_rls_policy(e):
                                st.error(
                                    "O Supabase bloqueou a gravação por RLS. Para salvar pela aplicação, configure `SUPABASE_SERVICE_ROLE_KEY` no Streamlit Cloud ou crie uma policy de insert/update para essa tabela."
                                )
                            else:
                                st.error(f"Erro ao salvar parametrizações pendentes das notas: {e}")
        elif fonte_documentos is not None and fonte_documentos.empty:
            st.warning("Não foi possível montar a lista de pendências porque a base de documentos está vazia.")
